import json
import os
import time
import random
import math

from openpyxl import load_workbook

from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize

import numpy as np
import pandas as pd

def find_nth(haystack, needle, n):
    """
    This function finds the index of the nth instance of a substring
    in a string
    """
    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start+len(needle))
        n -= 1
    return start

def clean_paper(paper):
    """
    This method takes a single paper and does all the rule-based text preprocessing.

    Parameters:
    ___________
    paper (str): The single paper to be preprocessed

    Returns:
    ________
    paper (str): The cleaned and preprocessed paper
    """
   # this series of statements cuts off the abstract/highlights/references/intro words
   # this method needs to be fixed, the elif sentences don't totally make sense
    if paper.lower().count('highlights') != 0:
        h_index = paper.lower().find('highlights')
        paper = paper[h_index + len('highlights'):]

    elif paper.lower().count('abstract') != 0:
        a_index = paper.lower().find('abstract')
        paper = paper[a_index + len('abstract'):]

    elif paper.lower().count('introduction') != 0:
        i_index = find_nth(paper.lower(),'introduction',2)
        paper = paper[i_index + len('introduction'):]

    else:
        pass

    r_index = paper.rfind('References')
    paper = paper[:r_index]

    return paper

def make_ner_sheet(journal_directory, retrieval_type='description', years='all', scramble=True,
                   pick_random=True, num_papers=1000, seed=42, pubs_per_sheet=100):
    """
    This function prepares text data to be labeled for NER training in excel spreadsheet.
    Attempts to pull papers from all years as equally as possible.

    Parameters:
        journal_directory (str, required): The absolute path to the journal directory

        retrieval_type (str, optional): Produce abstracts, or fulltexts. Use
            'abstract' for the former, and 'fulltext' for the latter.

        years (list-like, optional): List of years to collect papers from. If left
            as 'all', will grab papers from all years possible.

        scramble (bool, optional): Whether to scramble the papers year-wise throughout
            the excel spreadsheets, so the years from which the labeled data come from
            end up reasonably homogenous

        num_papers (int, optional): How many papers to put in the spreadsheets.
            Ceder et al stopped improving training accuracy after about 600-650
            abstracts worth of NER data
    """
    files = os.listdir(journal_directory)

    # find the .json file, containing all the publication data
    for filename in files:
        if filename.endswith('.json'):
            json_filename = filename
            break

    # find the jounal .json and make a dictionary from it
    paper_counter = 0
    os.chdir(journal_directory)
    with open(json_filename) as json_file:
        journal_dict = json.load(json_file)

    # calculate the publications per year we need
    if years != 'all':
        num_years = len(years)
        pubs_per_year = int(round(num_papers / num_years)) # round to nearest whole
    else:
        # the first layer of dictionaries in journal_dict corresponds to all the
        # years, so the number of keys will tell us how many pubs for each year
        num_years = len(journal_dict.keys())
        pubs_per_year = int(round(num_papers / num_years))

    if years == 'all':
        year_list = journal_dict.keys()
    else:
        year_list = years

    pubs = []
    pub_infos = []
    for year in year_list:


        year_dict = journal_dict[year]
        pubs_from_year = 0

        random.seed(seed)
        pub_idxs = random.sample(range(len(year_dict)), pubs_per_year)

        while pubs_from_year < pubs_per_year:
            pubs_from_year += 1

            # going to try/except because we don't know if pubs_per_year will
            # be greater than the number of publications in any of the years
            try:
                pub_tokens = []
                pub_idx = str(pub_idxs.pop()) # have to access the dictionary with strings. Yes, I suck

                # grab the relevant text corresponding to the random index
                if retrieval_type == 'description':
                    text = year_dict[pub_idx]['description'] # need to change this to 'abstract' in pyblio
                else:
                    text = year_dict[pub_idx]['fulltext']

                if text == None: # some abstracts and fulltexts just aren't there
                    continue
                text = clean_paper(text)
                sent_list = sent_tokenize(text)
                sent_endings = [] # will be built to track the index of sentence endings

                for i, sentence in enumerate(sent_list):
                    length = len(sentence.split())
                    sent_endings.append(length)
                    for word in sentence.split():
                        pub_tokens.append(word)

                pubs.append(pub_tokens)

                # edit sent_endings so it actually corresponds to tokens in pub_tokens
                # before this point, it's just a list of lengths, so doesn't work
                # for indexing
                for i, num in enumerate(sent_endings):
                    if i == 0:
                        pass
                    else:
                        sent_endings[i] = sent_endings[i] + sent_endings[i-1]
                sent_endings = np.array(sent_endings) - 1
                # at this point, sent_endings needs to be stored for later access
                # we will append them to info_tup, so they get carried with
                # the associated publication through the scramble


                info_tup = (year, pub_idx, year_dict[pub_idx]['doi'], year_dict[pub_idx]['pii'], sent_endings.tolist())
                pub_infos.append(info_tup)

            except:
                break

#     # grab the relevant text corresponding to the random index #########
#     if retrieval_type == 'description':
#         text = year_dict[pub_idx]['description'] # need to change this to 'abstract' in pyblio
#     else:
#         text = year_dict[pub_idx]['fulltext']
#
#     if text == None: # some abstracts and fulltexts just aren't there
#         continue
#     text = clean_paper(text)
#     for sentence in sent_tokenize(text):
#         for word in sentence.split():
#             pub_tokens.append(word)
#
#     pubs.append(pub_tokens)
#     info_tup = (year, pub_idx, year_dict[pub_idx]['doi'], year_dict[pub_idx]['pii'])
#     pub_infos.append(info_tup)
#
# except:
#     break ##########
    ########################### at this point, all pub tokens are made and stored
    np.random.seed(42)
    np.random.shuffle(pubs)

    np.random.seed(42)
    np.random.shuffle(pub_infos)

    longest = find_longest_paper(pubs)

    pubs_in_excel = 0
    sheet_number = 0
    pub_counter = 0
    master_endings_dict = {}


    while pubs_in_excel < len(pubs) - 1:

        pubs_in_sheet = 0
        dynamic_endings_dict = {}
        while pubs_in_sheet < pubs_per_sheet:
            try:
                data = pubs[pub_counter]
            except IndexError:
                break
            while len(data) < longest:
                data.append('')
            data = np.array(data)

            # create dataframe
            name   = ['' for x in range(len(data))]
            besio  = np.array(['' for x in range(len(data))])
            entity = besio
            mol_class = besio
            name[1] = pub_infos[pub_counter][0] # put the year in the second entry
            name[2] = pub_infos[pub_counter][1] # put the pub index within the year as third entry
            name[3] = pub_infos[pub_counter][2] # put the pub doi as fourth entry

            name = np.array(name) # now turn it into np array

            columns = ['name', 'tokens', 'BESIO', 'entity', 'mol_class']

            df = pd.DataFrame(np.array([name, data, besio, entity, mol_class]).transpose(), columns=columns)

            # write the damn thing to excel in the propper column
            sheet_name = json_filename[:-5]
            filename = f'{sheet_name}_{sheet_number}.xlsx'
            append_df_to_excel(filename, df, sheet_name=f'Sheet1', startcol = 6 * pubs_in_sheet)

            # append the sentence_endings lists into the dict
            dynamic_endings_dict[pubs_in_sheet] = pub_infos[pub_counter][4]

            pubs_in_sheet += 1
            pubs_in_excel += 1
            pub_counter += 1

        sheet_number += 1
        master_endings_dict[filename] = dynamic_endings_dict

    with open('sentence_endings.json', 'w') as fp:
        json.dump(master_endings_dict, fp)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=0, startcol=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...

      startcol : upper left cell column to dump data frame.


      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    print('On dataframe ', startcol/6)
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = writer.book[sheet_name].max_col

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass


    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()


def find_longest_paper(pubs):
    """
    This function finds the longest paper in a year_dict, in terms of how many
    tokens are in the paper.

    Parameters:
        pubs (list-like, required): The year_dict to be searched

    Returns:
        longest (int): The length of the longest paper in the year dict
    """
    longest = 0
    for paper in pubs:
        if len(paper) > longest:
            longest = len(paper)

    return longest


def make_journal_training_data(path):
    """
    This function takes in a labeled NER sheet, and finds the columns that have been labeled.

    Parameters:
        path (str, required): The absolute path to a directory containing the labeled
            NER training data for a SINGLE journal, and the sentence_endings .json.

    Returns:
        list: A list of tuples, where each tuple contains the x and y NER training data
    """
    filenames = os.listdir(path)
    sheets = [file for file in filenames if file.endswith('xlsx')]
    for filename in filenames:
        if filename.endswith('.json'):
            json_filename = filename
            break

    with open(json_filename, 'r') as fp:
        endings = json.load(fp) # the dictionary of sentence ending indecies
                                # for each text in each sheet
    training_data = []
    for sheet in sheets:
        sheet_df = pd.read_excel(sheet)
        endings_dict = endings[sheet]   # select the sentence end indices corresponding

        data = extract_xy(sheet_df, endings_dict)
        training_data.append(data)

    return training_data


def extract_xy(df, endings_dict):
    """
    This method extracts and correctly aranges the NER training x-values (tokens)
    and y-values (BESIO labels) from a pandas dataframe containing labeled NER
    data

    Parameters:
        df (pandas DataFrame, required): Dataframe loaded via pd.read_excel() on
            a labeled NER dataset

        endings_dict (dictionary, required): Dictionary containing the indicies
            where each sentence in each line of tokens ends.

    Returns:
        list: List of tuples, containing the x,y pairs
    """
    labeled = []
    columns = df.columns

    new_df = pd.DataFrame()
    training_y_iterators = []   # will contain a list of iterators
    training_x = []             # will contain lists of tokens, one for each text
    frame_indicies = []         # will contain indicies of dataframes that actually
                                # had NER labels in them

    for idx, column in enumerate(columns):
        # skip the columns that say Unnamed.x
        if column.startswith('Unnamed'):
            if idx == 0:
                frame_ticker = 0
            else:
                frame_ticker += 1
            continue
        else:
            pass

        # find every column that starts with 'name'
        if column.startswith('name'):

            # check if the entry in 'name' cell is a str
            if isinstance(df[column][0], str):
                # put the name column in new_df
                new_df[column] = df[column]

                tokens = df[columns[idx + 1]].values
                training_x.append(tokens)
                new_df[columns[idx + 1]] = tokens # put the tokens in new_df

                # append this successful index into idxs so we know which df in
                # the sheet this is
                frame_indicies.append(frame_ticker)

                # assume there is at least one label column after
                # start at 2 because first label column is +2 ahead of 'name' col
                nextcol = columns[idx + 2]

                # also start this at 2, increment in while-loop to look further into frame
                fwd_idx = 2
                master_labels = []
                while not (nextcol.startswith('Unnamed')):
                    if nextcol.startswith('BESIO'):
                        labels = df[nextcol].replace(np.nan, 'O')
                    else:
                        labels = df[nextcol].replace(np.nan, '')

                    # putting the labels into clean dataframe
                    new_df[nextcol] = labels
                    fwd_idx += 1
                    nextcol = columns[idx + fwd_idx]

                    # also putting the labels into a list to be concatenated
                    master_labels.append(labels.tolist())

                ###### FIX THIS NEXT LINE. IT IS NOT GENERALIZABLE #########
                zipper = zip(master_labels[0], master_labels[1], master_labels[2])
                training_y_iterators.append(zipper)

    data = arange_xy(training_x, training_y_iterators, endings_dict, frame_indicies)

    return data # used to return new_df

def arange_xy(training_x, training_y_iterators, endings_dict, frame_indicies):
    """
    This functions properly aranges the xy training data contained in training_x and
    training_y_iterators.

    Parameters:
        training_x (list, required): List of lists. Each nested list contains all
            the tokens from a single NER text (abstract or fulltext)

        training_y_iterators (list, required): List of iterators. The ith iterator
            corresponds to the ith token list in training_x. Each iterator iterates
            through one tuple for every token. Every tuple contains the label

        endings_dict (dict, required): Dictionary containing lists of token indecies
            where sentences end.

        frame_indicies (list, required): A list whose ith entry corresponds to the index
            of the ith training_x and ith trainin_y entry. Used to get the correct
            sentence endings list from endings_dict.

    Returns:
        list: list of tuples. Each tuple[0] is a training x-value, and each tuple[1]
            is a training y-value.
    """
    data = []
    for idx, tokens in enumerate(training_x):
        frame_number = str(frame_indicies[idx])         # get the correct frame number
        sentence_endings = endings_dict[frame_number]   # get the indicies of the endings of each sentence
        label_iterator = training_y_iterators[idx]      # get the corresponding iterator

        labels = build_labels(label_iterator)

        # rebuild the list of sentences
        sentences, label_lists = from_endings(tokens.tolist(), labels, sentence_endings)
        for idx2, sentence in enumerate(sentences):
            data.append((sentence, label_lists[idx2]))

    return data

def from_endings(tokens, labels, endings):
    """
    This function reconstructs the original sentences (and their corresponding
    labels) from a token list.

    Parameters:
        tokens (list, required): List of tokens to be reconstructed into sentences

        labels (list, required): List of labels, with ith label corresponding to
            ith token

        endings (list, required): List of indicies where sentences end in the token
            list.
    """
    sentences = []
    label_lists = []
    for i, ending in enumerate(endings):
        if i == 0:
            sentence = tokens[:ending + 1]
            sentence_labels = labels[:ending + 1]
            sentences.append(sentence)
            label_lists.append(sentence_labels)

        elif i != len(endings) - 1: # this would be the last ending in endings
            previous_ending = endings[i-1]
            start = previous_ending + 1

            sentence = tokens[start: ending + 1 ]
            sentence_labels = labels[start: ending + 1 ]
            sentences.append(sentence)
            label_lists.append(sentence_labels)

        else: # we are at the last sentence
            previous_ending = endings[i-1]
            start = previous_ending + 1

            sentence = tokens[start:]
            sentence_labels = labels[start:]
            sentences.append(sentence)
            label_lists.append(sentence_labels)

    return sentences, label_lists

def build_labels(label_iterator):
    """
    This function constructs a list of BESIO labels from an iterator over the
    label tuples.

    Parameters:
        label_iterator (iterator, required): Iterates through the label tuples
            corresponding to each token in a NER-labeled token list

    Returns:
        list: List of strings, each string is the concatenated BESIO label for
            a token
    """
    labels = []
    for entry in label_iterator:

        for i, tag in enumerate(entry):
            if i == 0:
                if tag != '':           # theoretically this is redundant if the
                    full_tag = tag      # BESIO column comes first
                else:
                    pass
            else:
                if tag != '':
                    full_tag = '-'.join((full_tag,tag))
                else:
                    pass
        labels.append(full_tag)
    return labels

def make_all_training_data(folder_path):
    """
    This function collects all the labeled NER training data from a folder holding
    any number of ner_sheet folders from different journals.

    Parameters:
        folder_path (str, required): The path to a folder containing other folders.
            Those "other folders" correspond to different journals, and contain
            filled out excel spreadsheets with NER training data.

    Returns:
        array: array of tuples. Each tuple contains one list of tokens, and the
            corresponding NER labels for each token.
    """
    # ensure we start with the same format every time
    if not folder_path.endswith('/'):
        folder_path += '/'

    ner_folders = os.listdir(folder_path)

    # iter through all NER data folders corresponding to a single journal
    for journal in ner_folders:
        journal_path = folder_path + journal + '/'

        for sheet in journal_path:

            if sheet.endswith('xlsx'): # don't want to open google sheets
                df = pd.read_excel(sheet)
                clean_df = find_labeled(df)
            else:
                pass


# def label_main():
#     """
#     This is the main method for the paper label exection.
#     """
#     carbon_path = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Carbon'
#     j_in_bio = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Journal_of_Inorganic_Biochemistry'
#     j_o_metallic = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Journal_of_Organometallic_Chemistry'
#     #make_ner_sheet(carbon_path, num_papers = 300, pubs_per_sheet = 50)
#     make_ner_sheet(j_in_bio, num_papers = 300, pubs_per_sheet = 50)
#     make_ner_sheet(j_o_metallic, num_papers = 300, pubs_per_sheet = 50)
#
# label_main()
