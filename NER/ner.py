import json
import os
import re
import time
import random
import math
import collections

from openpyxl import load_workbook

import pubchempy as pcp
from chemdataextractor import Document

from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize
# from mat2vec.mat2vec.processing.process import MaterialsTextProcessor
from mat2vec.processing.process import MaterialsTextProcessor

import numpy as np
import pandas as pd

# if debug is True, print statements occure to help identify problem sheet
debug = False

def find_nth(haystack, needle, n):
    """
    This function finds the index of the nth instance of a substring
    in a string
    """
    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start+len(needle))
        n -= 1
    return start

######## TEXT PROCESSING ##########

def clean_paper(paper):
    """
    This method takes a single paper and does all the rule-based text preprocessing.

    Parameters:
    ___________
    paper (str): The single paper to be preprocessed

    Returns:
    ________
    paper (str): The cleaned and preprocessed paper
    """
   # this series of statements cuts off the abstract/highlights/references/intro words
   # this method needs to be fixed, the elif sentences don't totally make sense
    if paper.lower().count('highlights') != 0:
        h_index = paper.lower().find('highlights')
        paper = paper[h_index + len('highlights'):]

    elif paper.lower().count('abstract') != 0:
        a_index = paper.lower().find('abstract')
        paper = paper[a_index + len('abstract'):]

    elif paper.lower().count('introduction') != 0:
        i_index = find_nth(paper.lower(),'introduction',2)
        paper = paper[i_index + len('introduction'):]

    else:
        pass

    r_index = paper.rfind('References')
    paper = paper[:r_index]

    return paper

def replace_with_space(text, rem_char):
    rem_idxs = []
    for i, char in enumerate(text):
        if char == rem_char:
            rem_idxs.append(i)
    for idx in rem_idxs:
        text = text[:idx] + ' ' + text[idx+1:]
    text = ' '.join(text.split())
    return text

def clean_highlights(text):
    """
    This method takes a CI abstract and does all the rule-based text preprocessing.

    Parameters:
    ___________
    paper (str): The single paper to be preprocessed

    Returns:
    ________
    paper (str): The cleaned and preprocessed paper
    """
    ### Check for highlights term
    if text.lower().count('highlights') != 0:

        ### Logic if highlights term leads the abstract
        if text.lower()[:10] == 'highlights':
            text = text[11:]
            a_index = text.lower().find('abstract')

            highlights_text = text[:a_index]
            abstract_text = text[a_index+len('abstract'):]

            if '•' in highlights_text:
                highlights_text = replace_with_space(highlights_text, '•')
            elif '?' in highlights_text:
                highlights_text = replace_with_space(highlights_text, '?')
            else:
                pass
            full_text = highlights_text + ' ' + abstract_text
            return full_text

        ### Logic if highlights term trails the abstract
        else:
            h_index = text.lower().find('highlights')
            highlights_text = text[h_index+len('highlights'):]
            if '►' in highlights_text:
                highlights_text = replace_with_space(highlights_text, '►')
                abstract_text = text[:h_index]
                full_text = abstract_text + ' ' + highlights_text
                return full_text
            elif '▸' in highlights_text:
                highlights_text = replace_with_space(highlights_text, '▸')
                abstract_text = text[:h_index]
                full_text = abstract_text + ' ' + highlights_text
                return full_text
            else:
                return text

    else:
        return text

def clean_leading_abstract(text):
    if text.lower()[:8] == 'abstract':
        text = text[8:]
        return text

    else:
        return text

def clean_graphics_text(text):
    if 'graphical abstract' in text.lower():
        g_index = text.lower().find('graphical abstract')
        text = text[:g_index]
        return text
    else:
        return text

def clean_CI_abstract(abstract):
    text = clean_highlights(abstract)
    text = clean_leading_abstract(text)
    text = clean_graphics_text(text)
    return text

def clean_Elsevier_abstract(abstract):
    abstract = abstract.split('\n')
    info = []
    for line in abstract:
        line = line.strip()
        if line != '':
            info.append(line)
    if len(info) == 2:
        abstract_type = info[0]
        clean_abstract = info[1]
    elif len(info) == 1:
        if info[0].split()[0].lower() == 'abstract':
            abstract_type = 'Abstract'
            clean_abstract = ' '.join(info[0].split()[1:])
        elif info[0].split()[0].lower() == 'summary':
            abstract_type = 'Summary'
            clean_abstract = ' '.join(info[0].split()[1:])
        elif 'objective' in info[0].split()[0].lower():
            abstract_type = 'Objective'
            clean_abstract = ' '.join(info[0].split()[1:])
        else:
            abstract_type = ''
            clean_abstract = info[0]
    else:
        info_lower = [x.lower() for x in info]
        section_titles = ['introduction',
                          'purpose',
                          'background',
                          'scope and approach',
                          'objective',
                          'objectives',
                          'materials and methods',
                          'results',
                          'conclusion',
                          'conclusions',
                          'key findings',
                          'key findings and conclusions',
                          'methodology',
                          'methods',
                          'study design',
                          'clinical implications']
        sectioned = False
        for section_title in section_titles:
            if section_title in info_lower:
                sectioned = True
        if sectioned:
            if info[0].lower() == 'abstract':
                abstract_type = 'Abstract'
                text = []
                for entry in info[1:]:
                    if entry.lower() in section_titles:
                        pass
                    else:
                        text.append(entry)
                clean_abstract = ' '.join(text)
            elif info[0].lower() == 'summary':
                abstract_type = 'Summary'
                text = []
                for entry in info[1::]:
                    if entry.lower() in section_titles:
                        pass
                    else:
                        text.append(entry)
                clean_abstract = ' '.join(text)
            else:
                abstract_type = ''
                text = []
                for entry in info:
                    if entry.lower() in section_titles:
                        pass
                    else:
                        text.append(entry)
                clean_abstract = ' '.join(text)
        else:
            if info[0].lower() == 'abstract' or info[0].lower() == 'absract' or info[0].lower() == 'abstact' or info[0].lower() == 'abstractt':
                abstract_type = 'Abstract'
                clean_abstract = ' '.join(info[1:])
            elif info[0].lower() == 'summary' or info[0].lower() == 'publisher summary' or info[0].lower() == '1. summary':
                abstract_type = 'Summary'
                clean_abstract = ' '.join(info[1:])
            elif info[0] == 'This article has been retracted: please see Elsevier Policy on Article Withdrawal (https://www.elsevier.com/about/our-business/policies/article-withdrawal).':
                abstract_type = 'Retracted'
                clean_abstract = 'Retracted'
            else:
                abstract_type = ''
                clean_abstract = ' '.join(info)
    return clean_abstract, abstract_type

def remove_abbreviations(abstract):
    doc = Document(abstract)
    abbvs = doc.abbreviation_definitions
    cems = doc.cems
    if len(abbvs) > 0:
        abbv_dict = {}
        for abbv in abbvs:
            cem_starts = []
            cem_ends = []
            if abbv[-1] is not None:
                abbv_dict[abbv[0][0]] = [' '.join(abbv[1])]
                for cem in cems:
                    if cem.text == abbv[0][0]:
                        cem_starts.append(cem.start)
                        cem_ends.append(cem.end)
                if len(cem_starts) > 0:
                    low_idx = cem_starts[np.argmin(cem_starts)]
                else:
                    low_idx = 0
                abbv_dict[abbv[0][0]].append(low_idx)
        abbv_dict = {k: v for k, v in sorted(abbv_dict.items(), key=lambda item: item[1][1])}
        index_change = 0
        for abbv in abbv_dict.keys():
            non_abbv = abbv_dict[abbv][0]
            if abbv_dict[abbv][1] != 0:
                replacement_delta = len(non_abbv) - len(abbv)
                cem_starts = []
                cem_ends = []
                for cem in cems:
                    if cem.text == abbv:
                        cem_starts.append(cem.start)
                        cem_ends.append(cem.end)
                if len(cem_starts) == 1:
                    if abstract[cem_starts[0]+index_change-1]+abstract[cem_ends[0]+index_change] == '()':
                        abstract = abstract[:cem_starts[0]-2+index_change] + abstract[cem_ends[0]+1+index_change:]
                        index_change += cem_starts[0] - cem_ends[0] - 3
                    else:
                        pass
                else:
                    low_idx = np.argmin(cem_starts)
                    cem_start_low = cem_starts[low_idx]
                    cem_end_low = cem_ends[low_idx]
                    if abstract[cem_start_low+index_change-1]+abstract[cem_end_low+index_change] == '()':
                        abstract = abstract[:cem_start_low-2+index_change] + abstract[cem_end_low+1+index_change:]
                        index_change += cem_start_low - cem_end_low - 3
                    else:
                        pass
                abstract = re.sub(r'([\s]){}([.,;\s]|$)'.format(abbv), r' {}\2'.format(non_abbv), abstract)
            else:
                pass
    return abstract

def normalize_elements(abstract):
    ELEMENTS = ["H", "He", "Li", "Be", "B", "C", "N", "O", "F", "Ne", "Na", "Mg", "Al", "Si", "P", "S", "Cl", "Ar", "K",
                "Ca", "Sc", "Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn", "Ga", "Ge", "As", "Se", "Br", "Kr",
                "Rb", "Sr", "Y", "Zr", "Nb", "Mo", "Tc", "Ru", "Rh", "Pd", "Ag", "Cd", "In", "Sn", "Sb", "Te", "I",
                "Xe", "Cs", "Ba", "La", "Ce", "Pr", "Nd", "Pm", "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", "Er", "Tm", "Yb",
                "Lu", "Hf", "Ta", "W", "Re", "Os", "Ir", "Pt", "Au", "Hg", "Tl", "Pb", "Bi", "Po", "At", "Rn", "Fr",
                "Ra", "Ac", "Th", "Pa", "U", "Np", "Pu", "Am", "Cm", "Bk", "Cf", "Es", "Fm", "Md", "No", "Lr", "Rf",
                "Db", "Sg", "Bh", "Hs", "Mt", "Ds", "Rg", "Cn", "Nh", "Fl", "Mc", "Lv", "Ts", "Og", "Uue"]

    ELEMENT_NAMES = ["hydrogen", "helium", "lithium", "beryllium", "boron", "carbon", "nitrogen", "oxygen", "fluorine",
                     "neon", "sodium", "magnesium", "aluminium", "silicon", "phosphorus", "sulfur", "chlorine", "argon",
                     "potassium", "calcium", "scandium", "titanium", "vanadium", "chromium", "manganese", "iron",
                     "cobalt", "nickel", "copper", "zinc", "gallium", "germanium", "arsenic", "selenium", "bromine",
                     "krypton", "rubidium", "strontium", "yttrium", "zirconium", "niobium", "molybdenum", "technetium",
                     "ruthenium", "rhodium", "palladium", "silver", "cadmium", "indium", "tin", "antimony", "tellurium",
                     "iodine", "xenon", "cesium", "barium", "lanthanum", "cerium", "praseodymium", "neodymium",
                     "promethium", "samarium", "europium", "gadolinium", "terbium", "dysprosium", "holmium", "erbium",
                     "thulium", "ytterbium", "lutetium", "hafnium", "tantalum", "tungsten", "rhenium", "osmium",
                     "iridium", "platinum", "gold", "mercury", "thallium", "lead", "bismuth", "polonium", "astatine",
                     "radon", "francium", "radium", "actinium", "thorium", "protactinium", "uranium", "neptunium",
                     "plutonium", "americium", "curium", "berkelium", "californium", "einsteinium", "fermium",
                     "mendelevium", "nobelium", "lawrencium", "rutherfordium", "dubnium", "seaborgium", "bohrium",
                     "hassium", "meitnerium", "darmstadtium", "roentgenium", "copernicium", "nihonium", "flerovium",
                     "moscovium", "livermorium", "tennessine", "oganesson", "ununennium"]
    element_dict = {}
    for element, name in zip(ELEMENTS, ELEMENT_NAMES):
        element_dict[element] = name
    element_dict['aluminium'] = 'aluminum'
    doc = Document(abstract)
    cems = doc.cems
    names = []
    starts = []
    ends = []
    for cem in cems:
        if cem.text in element_dict.keys():
            names.append(cem.text)
            starts.append(cem.start)
            ends.append(cem.end)
    names = np.array(names)
    starts = np.array(starts)
    ends = np.array(ends)
    sort = np.argsort(starts)
    names = names[sort]
    starts = starts[sort]
    ends = ends[sort]

    index_change = 0
    for name, start, end in zip(names, starts, ends):
        replace_name = element_dict[name]
        replace_delta = len(replace_name) - len(name)
        abstract = abstract[:start+index_change] + replace_name + abstract[end+index_change:]
        index_change += replace_delta
    return abstract

def remove_copywrite_CI(abstract):
    """
    This function takes an abstract and removes the copywrite information followed by the Elsevier
    text and publication year and returns a clean abstract.

    Parameters:
    abstract (str, required): The abstract which you want to clean
    """
    split= abstract.split()
    clean_abstracts = []

    if '©' in split:

        if split[0] != '©':
            index = split.index('©')
            del split[index:]
            clean_abstract = ' '.join(split)
            clean_abstracts.append(clean_abstract)


        elif split[0] == '©':
            if 'B.V.' in split:

                new_idx = split.index('B.V.')
                del split[0:new_idx+1]
                clean_abstract = ' '.join(split)
                clean_abstracts.append(clean_abstract)

            elif 'B.V..' in split:
                new_idxs = split.index('B.V..')
                del split[0:new_idxs+1]
                clean_abstract = ' '.join(split)
                clean_abstracts.append(clean_abstract)

            else:
    #             print(split)
                del split[0:2]
                clean_abstract = ' '.join(split)
                clean_abstracts.append(clean_abstracts)


    else:
        clean_abstract = ' '.join(split)
        clean_abstracts.append(clean_abstract)

    return clean_abstracts
    
def find_all_unique_entities(abstracts):
    entities = []
    for i, abstract in enumerate(abstracts):
        if i % 10 == 0:
            print('{} %'.format(round(i / len(abstracts) * 100, 3)))
        doc = Document(abstract)
        for j in range(len(doc.records)):
            try:
                entities.append(doc.records[j].serialize()['names'][0])
            except:
                pass
    unique_entities = list(set(entities))
    return unique_entities

def build_pubchem_synonym_dict(abstracts):
    entity_to_cid = {}
    cid_to_synonyms = {}
    for i, abstract in enumerate(abstracts):
        if i % 100 == 0:
            print('{} %'.format(round(i / len(abstracts) * 100, 2)))
        # Gather All Named Entities
        entities = []
        doc = Document(abstract)
        for j in range(len(doc.records)):
            try:
                entities.append(doc.records[j].serialize()['names'][0])
            except:
                pass

        # Gather Synonyms for Each CID
        for entity in entities:
            if entity.lower() in entity_to_cid.keys():
                pass
            else:
                c = pcp.get_compounds(entity, 'name')
                if len(c) >= 1:
                    try:
                        c = c[0]
                        cid = str(c.cid)
                        entity_to_cid[entity.lower()] = cid
                        if cid not in cid_to_synonyms.keys():
                            cid_to_synonyms[cid] = [entity]
                        else:
                            cid_to_synonyms[cid].append(entity)
                    except TimeoutError:
                        pass

    # Build Lookup Table for Each Named Entity With Synonyms
    lookup_dict = {}
    for entity, cid in entity_to_cid.items():
        lookup_dict[entity] = cid_to_synonyms[cid][0]

    return lookup_dict, entity_to_cid, cid_to_synonyms

def normalize_synonyms(abstract, lookup_dict):
    doc = Document(abstract)
    cems = doc.cems
    names = []
    starts = []
    ends = []
    for cem in cems:
        if cem.text.lower() in lookup_dict.keys():
            names.append(cem.text.lower())
            starts.append(cem.start)
            ends.append(cem.end)
    names = np.array(names)
    starts = np.array(starts)
    ends = np.array(ends)
    sort = np.argsort(starts)
    names = names[sort]
    starts = starts[sort]
    ends = ends[sort]

    index_change = 0
    for name, start, end in zip(names, starts, ends):
        replace_name = lookup_dict[name]
        replace_delta = len(replace_name) - len(name)
        abstract = abstract[:start+index_change] + replace_name + abstract[end+index_change:]
        index_change += replace_delta
    return abstract

######### LABELING #########

def make_ner_sheet(journal_directory, retrieval_type='description', years='all', scramble=True,
                   pick_random=True, num_papers=1000, seed=42, pubs_per_sheet=100):
    """
    This function prepares text data to be labeled for NER training in excel spreadsheet.
    Attempts to pull papers from all years as equally as possible.

    Parameters:
        journal_directory (str, required): The absolute path to the journal directory

        retrieval_type (str, optional): Produce abstracts, or fulltexts. Use
            'abstract' for the former, and 'fulltext' for the latter.

        years (list-like, optional): List of years to collect papers from. If left
            as 'all', will grab papers from all years possible.

        scramble (bool, optional): Whether to scramble the papers year-wise throughout
            the excel spreadsheets, so the years from which the labeled data come from
            end up reasonably homogenous

        num_papers (int, optional): How many papers to put in the spreadsheets.
            Ceder et al stopped improving training accuracy after about 600-650
            abstracts worth of NER data
    """
    files = os.listdir(journal_directory)

    # find the .json file, containing all the publication data
    for filename in files:
        if filename.endswith('.json'):
            json_filename = filename
            break

    # find the jounal .json and make a dictionary from it
    paper_counter = 0
    os.chdir(journal_directory)
    with open(json_filename) as json_file:
        journal_dict = json.load(json_file)

    # calculate the publications per year we need
    if years != 'all':
        num_years = len(years)
        pubs_per_year = int(round(num_papers / num_years)) # round to nearest whole
    else:
        # the first layer of dictionaries in journal_dict corresponds to all the
        # years, so the number of keys will tell us how many pubs for each year
        num_years = len(journal_dict.keys())
        pubs_per_year = int(round(num_papers / num_years))

    if years == 'all':
        year_list = journal_dict.keys()
    else:
        year_list = years

    pubs = []
    pub_infos = []
    for year in year_list:


        year_dict = journal_dict[year]
        pubs_from_year = 0

        random.seed(seed)
        pub_idxs = random.sample(range(len(year_dict)), pubs_per_year)

        while pubs_from_year < pubs_per_year:
            pubs_from_year += 1

            # going to try/except because we don't know if pubs_per_year will
            # be greater than the number of publications in any of the years
            try:
                pub_tokens = []
                pub_idx = str(pub_idxs.pop()) # have to access the dictionary with strings. Yes, I suck

                # grab the relevant text corresponding to the random index
                if retrieval_type == 'description':
                    text = year_dict[pub_idx]['description'] # need to change this to 'abstract' in pyblio
                else:
                    text = year_dict[pub_idx]['fulltext']

                if text == None: # some abstracts and fulltexts just aren't there
                    continue
                text = clean_paper(text)
                sent_list = sent_tokenize(text)
                sent_endings = [] # will be built to track the index of sentence endings

                for i, sentence in enumerate(sent_list):
                    length = len(sentence.split())
                    sent_endings.append(length)
                    for word in sentence.split():
                        pub_tokens.append(word)

                pubs.append(pub_tokens)

                # edit sent_endings so it actually corresponds to tokens in pub_tokens
                # before this point, it's just a list of lengths, so doesn't work
                # for indexing
                for i, num in enumerate(sent_endings):
                    if i == 0:
                        pass
                    else:
                        sent_endings[i] = sent_endings[i] + sent_endings[i-1]
                sent_endings = np.array(sent_endings) - 1
                # at this point, sent_endings needs to be stored for later access
                # we will append them to info_tup, so they get carried with
                # the associated publication through the scramble


                info_tup = (year, pub_idx, year_dict[pub_idx]['doi'], year_dict[pub_idx]['pii'], sent_endings.tolist())
                pub_infos.append(info_tup)

            except:
                break

    np.random.seed(42)
    np.random.shuffle(pubs)

    np.random.seed(42)
    np.random.shuffle(pub_infos)

    longest = find_longest_paper(pubs)

    pubs_in_excel = 0
    sheet_number = 0
    pub_counter = 0
    master_endings_dict = {}


    while pubs_in_excel < len(pubs) - 1:

        pubs_in_sheet = 0
        dynamic_endings_dict = {}
        while pubs_in_sheet < pubs_per_sheet:
            try:
                data = pubs[pub_counter]
            except IndexError:
                break
            while len(data) < longest:
                data.append('')
            data = np.array(data)

            # create dataframe
            name   = ['' for x in range(len(data))]
            besio  = np.array(['' for x in range(len(data))])
            entity = besio
            mol_class = besio
            name[1] = pub_infos[pub_counter][0] # put the year in the second entry
            name[2] = pub_infos[pub_counter][1] # put the pub index within the year as third entry
            name[3] = pub_infos[pub_counter][2] # put the pub doi as fourth entry

            name = np.array(name) # now turn it into np array

            columns = ['name', 'tokens', 'BESIO', 'entity', 'mol_class']

            df = pd.DataFrame(np.array([name, data, besio, entity, mol_class]).transpose(), columns=columns)

            # write the damn thing to excel in the propper column
            sheet_name = json_filename[:-5]
            filename = f'{sheet_name}_{sheet_number}.xlsx'
            append_df_to_excel(filename, df, sheet_name=f'Sheet1', startcol = 6 * pubs_in_sheet)

            # append the sentence_endings lists into the dict
            dynamic_endings_dict[pubs_in_sheet] = pub_infos[pub_counter][4]

            pubs_in_sheet += 1
            pubs_in_excel += 1
            pub_counter += 1

        sheet_number += 1
        master_endings_dict[filename] = dynamic_endings_dict

    with open('sentence_endings.json', 'w') as fp:
        json.dump(master_endings_dict, fp)


def make_CI_sheet(abstract_file, num_abstracts, file_name='CI_ner_labeling', seed=42, pubs_per_sheet=50):

    ff = open(abstract_file, 'r')
    file = ff.readlines()

    abstracts = []
    piis = []
    dois = []
    abs_infos = []

    random.seed(seed)
    abs_idxs = random.sample(range(len(file)), num_abstracts)
    random_abs = []
    random_piis = []
    random_dois = []
    for i in abs_idxs:
        random_abs.append(file[i].split('SPLITHERE')[0])
        random_piis.append(file[i].split('SPLITHERE')[1])
        random_dois.append(file[i].split('SPLITHERE')[2].split('\n')[0])

    processor = MaterialsTextProcessor()
    for i, (text, pii, doi) in enumerate(zip(random_abs, random_piis, random_dois)):
        sent_list = processor.tokenize(text, keep_sentences=True)
        sent_endings = []
        abs_tokens = []
        char_tokens = []

        for i, sentence in enumerate(sent_list):
            sentence, mat_list = processor.process(sentence)
            length = len(sentence)

            sent_endings.append(length)
            for word in sentence:
                abs_tokens.append(word)
                for char in word:
                    char_tokens.append(char)
        abstracts.append(abs_tokens)
        piis.append(pii)
        dois.append(doi)

        char_tokens = list(set(char_tokens))

        for i, num in enumerate(sent_endings):
            if i == 0:
                pass
            else:
                sent_endings[i] = sent_endings[i] + sent_endings[i-1]
        sent_endings = np.array(sent_endings) - 1

        info_tup = (i, sent_endings.tolist())
        abs_infos.append(info_tup)


    pubs_in_excel = 0
    sheet_number = 0
    pub_counter = 0
    master_endings_dict = {}


    while pubs_in_excel < len(abstracts) - 1:

        pubs_in_sheet = 0
        dynamic_endings_dict = {}
        while pubs_in_sheet < pubs_per_sheet:
            try:
                data = abstracts[pub_counter]
                pii = piis[pub_counter]
                doi = dois[pub_counter]
            except IndexError:
                break

            data = np.array(data)

            # create dataframe
            name   = ['' for x in range(len(data))]
            besio  = np.array(['' for x in range(len(data))])
            entity = besio
            mol_class = besio

            name = np.array(name) # now turn it into np array

            columns = ['name', 'tokens', 'BESIO', 'entity', 'mol_class']

            df = pd.DataFrame(np.array([name, data, besio, entity, mol_class]).transpose(), columns=columns)
            df['name'][0] = 'enter your name'
            df['name'][1] = pii
            df['name'][2] = doi

            # write the damn thing to excel in the propper column
            sheet_name = file_name
            filename = f'{sheet_name}_{sheet_number}.xlsx'
            append_df_to_excel(filename, df, sheet_name=f'Sheet1', startcol = 6 * pubs_in_sheet)

            # append the sentence_endings lists into the dict
            dynamic_endings_dict[pubs_in_sheet] = abs_infos[pub_counter][1]

            pubs_in_sheet += 1
            pubs_in_excel += 1
            pub_counter += 1

        sheet_number += 1
        master_endings_dict[filename] = dynamic_endings_dict

    with open('sentence_endings.json', 'w') as fp:
        json.dump(master_endings_dict, fp)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=0, startcol=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...

      startcol : upper left cell column to dump data frame.


      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    print('On dataframe ', startcol/6)
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = writer.book[sheet_name].max_col

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass


    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()


def find_longest_paper(pubs):
    """
    This function finds the longest paper in a year_dict, in terms of how many
    tokens are in the paper.

    Parameters:
        pubs (list-like, required): The year_dict to be searched

    Returns:
        longest (int): The length of the longest paper in the year dict
    """
    longest = 0
    for paper in pubs:
        if len(paper) > longest:
            longest = len(paper)

    return longest


def make_journal_training_data(path):
    """
    This function takes in a labeled NER sheet, and finds the columns that have been labeled.

    Parameters:
        path (str, required): The absolute path to a directory containing the labeled
            NER training data for a SINGLE journal, and the sentence_endings .json.

    Returns:
        list: A list of tuples, where each tuple contains the x and y NER training data
    """
    if not path.endswith('/'):
        path += '/'

    filenames = os.listdir(path)
    sheets = [file for file in filenames if file.endswith('xlsx')]
    if debug:
        print('This is the sheet list ', sheets)

    for filename in filenames:
        if filename.endswith('.json'):
            json_filename = filename
            break

    with open(path + json_filename, 'r') as fp:
        endings = json.load(fp) # the dictionary of sentence ending indecies
                                # for each text in each sheet
    training_data = []
    for sheet in sheets:

        # create a fake dataframe just to get the columns
        if debug:
            print(path+sheet)
        fake_df = pd.read_excel(path + sheet)
        convertable_cols = fake_df.columns
        converters = {col:str for col in convertable_cols}

        # the real sheet df, but making sure pandas interprets all data as strings
        sheet_df = pd.read_excel(path + sheet, converters=converters)
        endings_dict = endings[sheet]   # select the sentence end indices corresponding

        # print('This is path ', path)
        # print('This is sheet ', sheet)
        data = extract_xy(sheet_df, endings_dict)
        if len(data) != 0:
            training_data.append(data)

    return training_data


def extract_xy(df, endings_dict):
    """
    This method extracts and correctly aranges the NER training x-values (tokens)
    and y-values (BESIO labels) from a pandas dataframe containing labeled NER
    data

    Parameters:
        df (pandas DataFrame, required): Dataframe loaded via pd.read_excel() on
            a labeled NER dataset

        endings_dict (dictionary, required): Dictionary containing the indicies
            where each sentence in each line of tokens ends.

    Returns:
        list: List of tuples, containing the x,y pairs
    """
    labeled = []
    columns = df.columns

    new_df = pd.DataFrame()
    training_y_iterators = []   # will contain a list of iterators
    training_x = []             # will contain lists of tokens, one for each text
    frame_indicies = []         # will contain indicies of dataframes that actually
                                # had NER labels in them

    for idx, column in enumerate(columns):
        #print('This is column ', column)
        # skip the columns that say Unnamed.x
        if column.startswith('Unnamed'):
            if idx == 0:
                frame_ticker = 0
            else:
                frame_ticker += 1
            continue
        else:
            pass

        # find every column that starts with 'name'
        if column.startswith('name'):

            # check if the entry in 'name' cell is a str
            if isinstance(df[column][0], str):
                # put the name column in new_df
                new_df[column] = df[column]

                tokens = df[columns[idx + 1]].values # this used to just be df[columns[idx + 1]].values
                training_x.append(tokens)
                new_df[columns[idx + 1]] = tokens # put the tokens in new_df

                # append this successful index into idxs so we know which df in
                # the sheet this is
                frame_indicies.append(frame_ticker)

                # assume there is at least one label column after
                # start at 2 because first label column is +2 ahead of 'name' col
                nextcol = columns[idx + 2]

                # also start this at 2, increment in while-loop to look further into frame
                fwd_idx = 2
                master_labels = []
                while not (nextcol.startswith('Unnamed')):
                    if nextcol.startswith('BESIO'):
                        labels = df[nextcol].replace(np.nan, 'O')
                    else:
                        labels = df[nextcol].replace(np.nan, '')

                    # putting the labels into clean dataframe
                    new_df[nextcol] = labels
                    fwd_idx += 1

                    nextcol = columns[idx + fwd_idx]

                    # also putting the labels into a list to be concatenated
                    master_labels.append(labels.tolist())

                ###### FIX THIS NEXT LINE. IT IS NOT GENERALIZABLE #########
                zipper = zip(master_labels[0], master_labels[1], master_labels[2])
                training_y_iterators.append(zipper)

    data = arange_xy(training_x, training_y_iterators, endings_dict, frame_indicies)

    return data # used to return new_df

def arange_xy(training_x, training_y_iterators, endings_dict, frame_indicies):
    """
    This functions properly aranges the xy training data contained in training_x and
    training_y_iterators.

    Parameters:
        training_x (list, required): List of lists. Each nested list contains all
            the tokens from a single NER text (abstract or fulltext)

        training_y_iterators (list, required): List of iterators. The ith iterator
            corresponds to the ith token list in training_x. Each iterator iterates
            through one tuple for every token. Every tuple contains the label

        endings_dict (dict, required): Dictionary containing lists of token indecies
            where sentences end.

        frame_indicies (list, required): A list whose ith entry corresponds to the index
            of the ith training_x and ith trainin_y entry. Used to get the correct
            sentence endings list from endings_dict.

    Returns:
        list: list of tuples. Each tuple[0] is a training x-value, and each tuple[1]
            is a training y-value.
    """
    data = []
    for idx, tokens in enumerate(training_x):
        frame_number = str(frame_indicies[idx])         # get the correct frame number
        sentence_endings = endings_dict[frame_number]   # get the indicies of the endings of each sentence
        label_iterator = training_y_iterators[idx]      # get the corresponding iterator

        labels = build_labels(label_iterator)

        # rebuild the list of sentences
        sentences, label_lists = from_endings(tokens.tolist(), labels, sentence_endings)
        for idx2, sentence in enumerate(sentences):
            data.append((sentence, label_lists[idx2]))

    return data

def from_endings(tokens, labels, endings):
    """
    This function reconstructs the original sentences (and their corresponding
    labels) from a token list.

    Parameters:
        tokens (list, required): List of tokens to be reconstructed into sentences

        labels (list, required): List of labels, with ith label corresponding to
            ith token

        endings (list, required): List of indicies where sentences end in the token
            list.
    """
    sentences = []
    label_lists = []
    for i, ending in enumerate(endings):
        if i == 0:
            sentence = tokens[:ending + 1]
            sentence_labels = labels[:ending + 1]
            sentences.append(sentence)
            label_lists.append(sentence_labels)

        elif i != len(endings) - 1: # this would be the last ending in endings
            previous_ending = endings[i-1]
            start = previous_ending + 1

            sentence = tokens[start: ending + 1 ]
            sentence_labels = labels[start: ending + 1 ]
            sentences.append(sentence)
            label_lists.append(sentence_labels)

        else: # we are at the last sentence
            try: # this works for sentences that aren't the longest sentence
                previous_ending = endings[i-1]
                start = previous_ending + 1

                sentence = tokens[start:ending+1]
                sentence_labels = labels[start:ending+1]
            except IndexError: # we are at the longest sentence,
                previous_ending = endings[i-1]
                start = previous_ending + 1

                sentence = tokens[start:]
                sentence_labels = labels[start:]
            sentences.append(sentence)
            label_lists.append(sentence_labels)

    return sentences, label_lists

def build_labels(label_iterator):
    """
    This function constructs a list of BESIO labels from an iterator over the
    label tuples.

    Parameters:
        label_iterator (iterator, required): Iterates through the label tuples
            corresponding to each token in a NER-labeled token list

    Returns:
        list: List of strings, each string is the concatenated BESIO label for
            a token
    """
    labels = []
    for entry in label_iterator:

        for i, tag in enumerate(entry):
            if i == 0:
                if tag != '':           # theoretically this is redundant if the
                    full_tag = clean_tag(tag)     # BESIO column comes first
                else:
                    pass
            else:
                if tag != '':
                    full_tag = '-'.join((full_tag, clean_tag(tag)))
                else:
                    pass

        # if for some reason the full tag ended up being 1 character
        # assume it should be an 'O'
        if len(full_tag) == 1:
            full_tag = 'O'

        labels.append(full_tag)
    return labels

def clean_tag(tag):
    """
    Function to clean up human tagging mistakes. Non generalizable function.

    Need to edit in future to make more generalizeable
    """
    tag = tag.upper().strip()
    if not tag.isalnum():
        tag = '<Unknown>'

    elif 'MO' in tag or 'OL' in tag:
        tag = 'MOL'

    elif 'P' in tag and 'R' in tag:
        tag = 'PRO'

    elif tag == 'EE':
        tag = 'E'

    elif tag == 'BB':
        tag = 'B'

    return tag



def make_all_training_data(folder_path):
    """
    This function collects all the labeled NER training data from a folder holding
    any number of ner_sheet folders from different journals.

    Parameters:
        folder_path (str, required): The path to a folder containing other folders.
            Those "other folders" correspond to different journals, and contain
            filled out excel spreadsheets with NER training data.

    Returns:
        array: array of tuples. Each tuple contains one list of tokens, and the
            corresponding NER labels for each token.
    """
    training_data = {}
    # ensure we start with the same format every time
    if not folder_path.endswith('/'):
        folder_path += '/'

    ner_folders = os.listdir(folder_path)

    # iter through all NER data folders corresponding to a single journal
    for journal in ner_folders:
        journal_path = folder_path + journal + '/'

        training_data[journal] = make_journal_training_data(journal_path)

    return training_data

class NerData():

    def __init__(self, path, w2v_model):
        """
        Parameters:
            path (str, required): path to a directory of sub-directories containig
                labeled NER data for different journals in .xlsx files

            w2v_model (gensim w2v model, required): trained word2vec model containing
                word embeddings
        """
        self.training_dict = make_all_training_data(path)
        self.w2v_model = w2v_model

        self.word_set, self.char_set, self.label_set = self.word_char_label_sets(self.training_dict, w2v_model)

        self.word2ix, self.ix2word = self.set2ix(self.word_set)
        self.char2ix, self.ix2char = self.set2ix(self.char_set)
        self.label2ix, self.ix2label = self.set2ix(self.label_set)
        self.list = self.tolist()


    def set2ix(self, set):
        """
        This method iterates through a set to create a dictionary of unique integer
        mappings for each element in the set.

        Parameters:
            set (python set): set of unique values to get mappings for

        Returns:
            dict: dictionary of integer mappings for each element in set
        """
        set_to_ix = {}
        for item in set:
            set_to_ix[item] = len(set_to_ix)

        ix_to_set = {set_to_ix[key]:key for key in set_to_ix}

        return set_to_ix, ix_to_set


    def word_char_label_sets(self, training_dict, w2v_model):
        """
        This method creates the sets of all possible words, characters, and labels present in a
        NER training data dictionary.

        Parameters:
            training_dict (dict, required): Training data dictionary from make_all_training_data

            w2v_model (gensim w2v model, required): The word2vec model that will be referenced
                for word embeddings and vocabulary
        """
        w2v_words = w2v_model.wv.vocab.keys() # they keys to the word embeddings
        label_set = []

        for journal_data in training_dict.keys():
            for sheet_data in training_dict[journal_data]:
                for xy in sheet_data:

                    sent_list = xy[0]
                    tags = xy[1]

                    for tag in set(tags):
                        if tag not in label_set:
                            label_set.append(tag)

        word_set = []
        char_set = []
        for word in w2v_words:
            word_set.append(word)
            for char in word:
                if char not in char_set:
                    char_set.append(char)

        word_set = set(word_set)
        char_set = set(char_set)
        label_set = set(label_set)

        return word_set, char_set, label_set

    def tolist(self):
        """
        This method creates one long list of training xy pairs.
        """
        master_list = []
        for journal_data in self.training_dict.keys():
            for sheet_data in self.training_dict[journal_data]:
                for xy in sheet_data:

                    master_list.append(xy)

        return master_list

# def label_main():
#     """
#     This is the main method for the paper label exection.
#     """
#     carbon_path = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Carbon'
#     j_in_bio = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Journal_of_Inorganic_Biochemistry'
#     j_o_metallic = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Journal_of_Organometallic_Chemistry'
#     #make_ner_sheet(carbon_path, num_papers = 300, pubs_per_sheet = 50)
#     make_ner_sheet(j_in_bio, num_papers = 300, pubs_per_sheet = 50)
#     make_ner_sheet(j_o_metallic, num_papers = 300, pubs_per_sheet = 50)
#
# label_main()
