"""
This script is for training word2vec on a corpus of text collected by the elsapy/pyblio modules
"""
import numpy as np
import pandas as pd
import gensim
import gensim, logging
from gensim.models import Word2Vec
import json
import os
import time
import random

from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize

def find_nth(haystack, needle, n):
    """
    This function finds the index of the nth instance of a substring
    in a string
    """
    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start+len(needle))
        n -= 1
    return start

def clean_paper(paper):
    """
    This method takes a single paper and does all the rule-based text preprocessing.

    Parameters:
    ___________
    paper (str): The single paper to be preprocessed

    Returns:
    ________
    paper (str): The cleaned and preprocessed paper
    """
   # this series of statements cuts off the abstract/highlights/references/intro words
   # this method needs to be fixed, the elif sentences don't totally make sense
    if paper.lower().count('highlights') != 0:
        h_index = paper.lower().find('highlights')
        paper = paper[h_index + len('highlights'):]

    elif paper.lower().count('abstract') != 0:
        a_index = paper.lower().find('abstract')
        paper = paper[a_index + len('abstract'):]

    elif paper.lower().count('introduction') != 0:
        i_index = find_nth(paper.lower(),'introduction',2)
        paper = paper[i_index + len('introduction'):]

    else:
        pass

    r_index = paper.rfind('References')
    paper = paper[:r_index]

    return paper


class Loader():
    """
    This class creates an iterator that can yield sentences from a given journal
    into gensim's Word2Vec.
    """

    def __init__(self, journal_directory, years='all', retrieval_type='fulltext', mode='w2v'):
        """
        This method creates an object, ready to iter through the tokens in a
        journal dictionary's 'fulltext' entry, or 'abstract' entry. This class
        can be used as a sentence generator to train word2vec, or to produce
        training data for labeling.

        Parameters:
            journal_directory (str, required): The absolute path to the journal
                .json file being used

            years (list-like, optional): List of years from which to pull publication
                data from. If left as 'all', all the years will be done.

            retrieval_type (str, optional): Produce abstracts, or fulltexts. Use
                'abstract' for the former, and 'fulltext' for the latter.
        """

        self.journal_directory = journal_directory
        self.years = years
        self.retrieval_type = retrieval_type

    def __iter__(self):
        for filename in os.listdir(self.journal_directory):
        # only open the json file in the directory for the journal
        # this for loop is super quick but there is probably a better way
        # just grab the json
            if '.json' in filename:
                paper_counter = 0
                word_counter = 0
                os.chdir(self.journal_directory)
                with open(filename) as json_file:
                    journal_dict = json.load(json_file)

                for year in journal_dict:

                    year_dict = journal_dict[year]

                    for paper_number in year_dict:
                        paper_dict = year_dict[paper_number]

                        if self.retrieval_type == 'fulltext':
                            text = paper_dict['fulltext']
                        else: # it was 'abstract'
                            text = paper_dict['abstract']

                        text = clean_paper(text)
                        paper_counter += 1
                        print('Papers is at ',paper_counter)
                        #print('counter is at ', paper_counter, ' papers.')

                        for sentence in sent_tokenize(text):
                            words = word_tokenize(sentence)
                            yield(words)


def make_ner_sheet(journal_directory, retrieval_type='description', years='all', scramble=True,
                   pick_random=True, num_papers=1000, seed=42, pubs_per_sheet=100):
    """
    This function prepares text data to be labeled for NER training in excel spreadsheet.
    Attempts to pull papers from all years as equally as possible.

    Parameters:
        journal_directory (str, required): The absolute path to the journal directory

        retrieval_type (str, optional): Produce abstracts, or fulltexts. Use
            'abstract' for the former, and 'fulltext' for the latter.

        years (list-like, optional): List of years to collect papers from. If left
            as 'all', will grab papers from all years possible.

        scramble (bool, optional): Whether to scramble the papers year-wise throughout
            the excel spreadsheets, so the years from which the labeled data come from
            end up reasonably homogenous

        num_papers (int, optional): How many papers to put in the spreadsheets.
            Ceder et al stopped improving training accuracy after about 600-650
            abstracts worth of NER data.
    """
    files = os.listdir(journal_directory)

    # find the .json file, containing all the publication data
    for filename in files:
        if filename.endswith('.json'):
            json_filename = filename
            break

    # find the jounal .json and make a dictionary from it
    paper_counter = 0
    os.chdir(journal_directory)
    with open(json_filename) as json_file:
        journal_dict = json.load(json_file)

    # calculate the publications per year we need
    if years != 'all':
        num_years = len(years)
        pubs_per_year = int(round(num_papers / num_years)) # round to nearest whole
    else:
        # the first layer of dictionaries in journal_dict corresponds to all the
        # years, so the number of keys will tell us how many pubs for each year
        num_years = len(journal_dict.keys())
        pubs_per_year = int(round(num_papers / num_years))

    if years == 'all':
        year_list = journal_dict.keys()
    else:
        year_list = years

    pubs = []
    pub_infos = []
    for year in year_list:


        year_dict = journal_dict[year]
        pubs_from_year = 0

        random.seed(seed)
        pub_idxs = random.sample(range(len(year_dict)), pubs_per_year)

        while pubs_from_year < pubs_per_year:
            pubs_from_year += 1

            # going to try/except because we don't know if pubs_per_year will
            # be greater than the number of publications in any of the years
            try:
                pub_tokens = []
                pub_idx = str(pub_idxs.pop()) # have to access the dictionary with strings. Yes, I suck

                # grab the relevant text corresponding to the random index
                if retrieval_type == 'description':
                    text = year_dict[pub_idx]['description'] # need to change this to 'abstract' in pyblio
                else:
                    text = year_dict[pub_idx]['fulltext']

                if text == None: # some abstracts and fulltexts just aren't there
                    continue
                text = clean_paper(text)
                for sentence in sent_tokenize(text):
                    for word in sentence.split():
                        pub_tokens.append(word)

                pubs.append(pub_tokens)
                info_tup = (year, pub_idx, year_dict[pub_idx]['doi'], year_dict[pub_idx]['pii'])
                pub_infos.append(info_tup)

            except:
                break

    longest = find_longest_paper(pubs)

    pubs_in_excel = 0
    sheet_number = 0
    pub_counter = 0


    while pubs_in_excel < len(pubs) - 1:

        pubs_in_sheet = 0
        while pubs_in_sheet < pubs_per_sheet:


            data = pubs[pub_counter]
            while len(data) < longest:
                data.append('')
            data = np.array(data)

            # create dataframe
            name   = ['' for x in range(len(data))]
            besio  = np.array(['' for x in range(len(data))])
            entity = np.array(['' for x in range(len(data))])

            name[1] = pub_infos[pub_counter][2] # put the doi in second entry
            name = np.array(name) # now turn it into np array

            columns = ['name', 'tokens', 'BESIO', 'entity (MOL/PRO)']

            df = pd.DataFrame(np.array([name, data, besio, entity]).transpose(), columns=columns)

            # write the damn thing to excel in the propper column
            filename = 'carbon_ner_labels.xlsx'
            append_df_to_excel(filename, sheet_name=f'Sheet{sheet_number}', startcol = 6 * pubs_in_sheet)

            pubs_in_sheet += 1
            pubs_in_excel += 1
            pub_counter += 1

        sheet_number += 1


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=0, startcol=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...

      startcol : upper left cell column to dump data frame.


      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = writer.book[sheet_name].max_col

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass


    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()




def find_longest_paper(pubs):
    """
    This function finds the longest paper in a year_dict, in terms of how many
    tokens are in the paper.

    Parameters:
        pubs (list-like, required): The year_dict to be searched

    Returns:
        longest (int): The length of the longest paper in the year dict
    """
    longest = 0
    for paper in pubs:
        if len(paper) > longest:
            longest = len(paper)

    return longest




def label_main():
    """
    This is the main method for the paper label exection.
    """
    carbon_path = '/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Carbon'
    make_ner_sheet(carbon_path, num_papers = 500)

label_main()




# def w2v_main():
#     """
#     This is the main method to be executed
#     """
#     generator = Loader('/gscratch/pfaendtner/dacj/nlp/fulltext_pOmOmOo/Carbon')
#     model = Word2Vec(generator, min_count=10, workers=1, size=200)
#     os.chdir('/gscratch/pfaendtner/dacj/nlp')
#     model.save('carbon.model')
#
# main()
